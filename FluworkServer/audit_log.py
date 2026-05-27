# audit_log.py
"""
Módulo centralizado de logging para trazabilidad
Usado por app.py, text_search.py, build_text_index.py
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
import json

LOG_PATH = Path("log_consultas.xlsx")

def registrar_log(usuario: str, endpoint: str, filtros: dict, n_result: int):
    """
    Registra operaciones en Excel (trazabilidad)

    Args:
        usuario: Identificador del usuario (ej: "anon", "system")
        endpoint: Endpoint o acción (ej: "/api/buscar", "build_index")
        filtros: Dict con parámetros de la operación
        n_result: Cantidad de resultados/fichas procesadas
    """
    try:
        if LOG_PATH.exists():
            wb = load_workbook(LOG_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["timestamp", "usuario", "endpoint", "filtros", "resultados"])

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            usuario,
            endpoint,
            json.dumps(filtros, ensure_ascii=False),
            n_result
        ])
        wb.save(LOG_PATH)
    except Exception as e:
        # No romper la app por problemas de log
        print(f"LOG ERROR: {e}")
